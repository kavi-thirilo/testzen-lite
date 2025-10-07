#!/usr/bin/env python3
"""
Excel utility functions for TestZen automation framework
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelManager:
    """Manages Excel file operations for test data"""

    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.df = None

    def load_test_data(self):
        """Load test data from Excel file"""
        try:
            # Try to load from TestCases sheet first, then from first non-Test_Summary sheet
            try:
                self.df = pd.read_excel(self.excel_file, sheet_name='TestCases')
            except:
                # If TestCases sheet doesn't exist, find the first sheet that's not Test_Summary
                excel_file_info = pd.ExcelFile(self.excel_file)
                sheet_name = None
                for sheet in excel_file_info.sheet_names:
                    if sheet != 'Test_Summary':
                        sheet_name = sheet
                        break
                if sheet_name:
                    self.df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
                else:
                    self.df = pd.read_excel(self.excel_file)
            
            print(f"[TZ] Loaded test data from {self.excel_file}")
            print(f"[TZ] Found {len(self.df)} test steps")
            
            # Clear Status and Result Message columns before starting test
            self.clear_status_columns()
            
            return True
        except Exception as e:
            print(f"[TZ] Failed to load Excel file: {e}")
            return False
    
    def clear_status_columns(self):
        """Clear Status and Result Message columns before test run"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Find status column with simplified column order: S.No | Description | Action | Locator Type | Locator Value | Input Data | Status | Result Message
            status_col = 7  # Column G (Status)
            result_col = 8  # Column H (Result Message)
            
            # Clear all status and result message cells (starting from row 2 to skip header)
            for row_num in range(2, ws.max_row + 1):
                # Clear status column
                ws.cell(row=row_num, column=status_col).value = ""
                ws.cell(row=row_num, column=status_col).fill = PatternFill(fill_type=None)
                
                # Clear result message column
                ws.cell(row=row_num, column=result_col).value = ""
            
            wb.save(self.excel_file)
            print("[TZ] Cleared previous test status from Excel")
            return True
        except Exception as e:
            print(f"[TZ] Failed to clear Excel status columns: {e}")
            return False

    def get_test_steps(self):
        """Get test steps as list of dictionaries"""
        if self.df is not None:
            return self.df.to_dict('records')
            return []

    def update_step_status(self, step_index, status, result_message=""):
        """Update step status in Excel"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Find status column with simplified column order: S.No | Description | Action | Locator Type | Locator Value | Input Data | Status | Result Message
            status_col = 7  # Column G (Status)
            result_col = 8  # Column H (Result Message)
            
            # Update status (step_index + 2 because of header row and 0-based index)
            row_num = step_index + 2
            ws.cell(row=row_num, column=status_col, value=status)
            
            # Color coding with more vibrant, Excel-standard colors
            if status == "PASSED":
                # Bright green background for passed tests
                ws.cell(row=row_num, column=status_col).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                ws.cell(row=row_num, column=status_col).font = Font(bold=True, color="000000")  # Black bold text
            elif status == "FAILED":
                # Bright red background for failed tests  
                ws.cell(row=row_num, column=status_col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=row_num, column=status_col).font = Font(bold=True, color="FFFFFF")  # White bold text
            
            # Add result message if provided
            if result_message:
                ws.cell(row=row_num, column=result_col, value=result_message)
            
            wb.save(self.excel_file)
            
        except Exception as e:
            print(f"[TZ] Error updating Excel status: {e}")
            
        # Also update dataframe if available
        if self.df is not None and step_index < len(self.df):
            if 'Status' not in self.df.columns:
                self.df['Status'] = ''
            if 'Result' not in self.df.columns:
                self.df['Result'] = ''
            self.df.at[step_index, 'Status'] = status
            self.df.at[step_index, 'Result'] = result_message

    def generate_test_summary(self, test_name=None, platform="ANDROID", execution_time=None, additional_info=None):
        """Generate a professionally formatted Test_Summary sheet"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            # Create or get Test_Summary sheet
            if 'Test_Summary' in wb.sheetnames:
                # Delete existing sheet and create new one
                wb.remove(wb['Test_Summary'])
            
            ws = wb.create_sheet('Test_Summary', 0)  # Insert as first sheet
            
            # Get test data from main sheet for analysis (find the sheet with test data)
            main_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name != 'Test_Summary':
                    main_sheet = wb[sheet_name]
                    break
            if main_sheet is None:
                main_sheet = wb.active
            total_steps = 0
            passed_steps = 0
            failed_steps = 0
            unexecuted_steps = 0
            
            # Analyze results from main sheet (assuming status is in column G)
            for row_num in range(2, main_sheet.max_row + 1):
                status_cell = main_sheet.cell(row=row_num, column=7)
                if status_cell.value:
                    total_steps += 1
                    if status_cell.value == "PASSED":
                        passed_steps += 1
                    elif status_cell.value == "FAILED":
                        failed_steps += 1
                else:
                    unexecuted_steps += 1
            
            # Calculate success rate
            success_rate = (passed_steps / total_steps * 100) if total_steps > 0 else 0
            overall_status = "PASSED" if failed_steps == 0 and total_steps > 0 else "FAILED" if failed_steps > 0 else "NOT_EXECUTED"
            
            # Define styles
            title_font = Font(name='Calibri', size=16, bold=True, color='2F5F8F')
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            content_font = Font(name='Calibri', size=11)
            bold_content_font = Font(name='Calibri', size=11, bold=True)
            
            header_fill = PatternFill(start_color='2F5F8F', end_color='2F5F8F', fill_type='solid')
            light_blue_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Row 1: Main Title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "TEST EXECUTION SUMMARY REPORT"
            title_cell.font = title_font
            title_cell.alignment = center_alignment
            ws.row_dimensions[1].height = 25
            
            # Row 3: Test Information Header
            ws.merge_cells('A3:F3')
            test_info_header = ws['A3']
            test_info_header.value = "TEST INFORMATION"
            test_info_header.font = header_font
            test_info_header.fill = header_fill
            test_info_header.alignment = center_alignment
            test_info_header.border = thin_border
            ws.row_dimensions[3].height = 20
            
            # Test Information Section (Rows 4-10)
            test_info = [
                ("Test Name:", test_name or os.path.basename(self.excel_file).replace('.xlsx', '')),
                ("Test Platform:", platform),
                ("Execution Date:", datetime.now().strftime("%Y-%m-%d")),
                ("Execution Time:", execution_time or datetime.now().strftime("%H:%M:%S")),
                ("Test Framework:", "TestZen Automation Framework"),
                ("Test Type:", "End-to-End Automated Test"),
                ("Environment:", additional_info.get('environment', 'Test Environment') if additional_info else 'Test Environment')
            ]
            
            for i, (label, value) in enumerate(test_info, start=4):
                # Label column (A)
                label_cell = ws.cell(row=i, column=1)
                label_cell.value = label
                label_cell.font = bold_content_font
                label_cell.fill = light_blue_fill
                label_cell.alignment = left_alignment
                label_cell.border = thin_border
                
                # Value columns (B-F merged)
                ws.merge_cells(f'B{i}:F{i}')
                value_cell = ws.cell(row=i, column=2)
                value_cell.value = value
                value_cell.font = content_font
                value_cell.alignment = left_alignment
                value_cell.border = thin_border
            
            # Row 12: Execution Results Header
            ws.merge_cells('A12:F12')
            results_header = ws['A12']
            results_header.value = "EXECUTION RESULTS"
            results_header.font = header_font
            results_header.fill = header_fill
            results_header.alignment = center_alignment
            results_header.border = thin_border
            ws.row_dimensions[12].height = 20
            
            # Results Summary Table (Rows 13-17)
            results_data = [
                ("Metric", "Count", "Percentage", "", "", "Status"),
                ("Total Test Steps", str(total_steps + unexecuted_steps), "100%", "", "", ""),
                ("Passed Steps", str(passed_steps), f"{(passed_steps/(total_steps + unexecuted_steps)*100):.1f}%" if (total_steps + unexecuted_steps) > 0 else "0%", "", "", "[PASS] PASSED" if passed_steps > 0 else ""),
                ("Failed Steps", str(failed_steps), f"{(failed_steps/(total_steps + unexecuted_steps)*100):.1f}%" if (total_steps + unexecuted_steps) > 0 else "0%", "", "", "[FAIL] FAILED" if failed_steps > 0 else ""),
                ("Unexecuted Steps", str(unexecuted_steps), f"{(unexecuted_steps/(total_steps + unexecuted_steps)*100):.1f}%" if (total_steps + unexecuted_steps) > 0 else "0%", "", "", "[PEND] PENDING" if unexecuted_steps > 0 else "")
            ]
            
            for i, row_data in enumerate(results_data, start=13):
                for j, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = center_alignment if j > 1 else left_alignment
                    
                    if i == 13:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                    elif i == 14:  # Total row
                        cell.font = bold_content_font
                        cell.fill = light_blue_fill
                    elif i == 15:  # Passed row
                        cell.font = content_font
                        if passed_steps > 0:
                            cell.fill = green_fill
                    elif i == 16:  # Failed row
                        cell.font = content_font
                        if failed_steps > 0:
                            cell.fill = red_fill
                    elif i == 17:  # Unexecuted row
                        cell.font = content_font
                        if unexecuted_steps > 0:
                            cell.fill = yellow_fill
            
            # Row 19: Overall Test Status
            ws.merge_cells('A19:F19')
            overall_cell = ws['A19']
            overall_cell.value = f"OVERALL TEST STATUS: {overall_status}"
            overall_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            overall_cell.alignment = center_alignment
            overall_cell.border = thin_border
            ws.row_dimensions[19].height = 25
            
            # Set status color
            if overall_status == "PASSED":
                overall_cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
            elif overall_status == "FAILED":
                overall_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                overall_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            
            # Row 21: Test Success Rate
            ws.merge_cells('A21:F21')
            success_cell = ws['A21']
            success_cell.value = f"SUCCESS RATE: {success_rate:.1f}%"
            success_cell.font = Font(name='Calibri', size=12, bold=True, color='2F5F8F')
            success_cell.alignment = center_alignment
            success_cell.fill = light_blue_fill
            success_cell.border = thin_border
            ws.row_dimensions[21].height = 20
            
            # Row 23: Additional Notes Header (if additional_info provided)
            if additional_info and additional_info.get('notes'):
                ws.merge_cells('A23:F23')
                notes_header = ws['A23']
                notes_header.value = "ADDITIONAL NOTES"
                notes_header.font = header_font
                notes_header.fill = header_fill
                notes_header.alignment = center_alignment
                notes_header.border = thin_border
                ws.row_dimensions[23].height = 20
                
                # Notes content
                ws.merge_cells('A24:F26')
                notes_cell = ws['A24']
                notes_cell.value = additional_info.get('notes', '')
                notes_cell.font = content_font
                notes_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                notes_cell.border = thin_border
                ws.row_dimensions[24].height = 60
            
            # Set column widths
            column_widths = [20, 15, 15, 10, 10, 20]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Set the main test sheet as active (not Test_Summary)
            if main_sheet:
                wb.active = main_sheet
            
            # Save workbook
            wb.save(self.excel_file)
            print(f"[TZ] Generated professional Test_Summary sheet in {self.excel_file}")
            return True
            
        except Exception as e:
            print(f"[TZ] Error generating Test_Summary sheet: {e}")
            return False

